import tkinter as tk
import subprocess
from datetime import datetime
import signal
import time
from pathlib import Path
from rtsp_config import rtsp_url
from define_config import VALID_DEFINES, CURRENT_DEFINE
from openpyxl import Workbook, load_workbook
import threading
import json
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
import socket

recording = False
ffmpeg_process = None
current_video_path = None

pending_barcode = None
current_barcode = None

BASE_DIR = Path(__file__).resolve().parent
VIDEO_ROOT = BASE_DIR / "video"
EXCEL_ROOT = BASE_DIR / "excel"


def build_daily_dir(root_dir, dt):
    return root_dir / dt.strftime("%Y") / dt.strftime("%m") / dt.strftime("%d")


def get_excel_path(dt):
    daily_dir = build_daily_dir(EXCEL_ROOT, dt)
    daily_dir.mkdir(parents=True, exist_ok=True)
    return daily_dir / f"{dt.strftime('%Y%m%d')}.xlsx"


def ensure_excel_file(excel_path):
    if excel_path.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "log"
    ws.append(["Thoi gian dong", "Ma vach", "Duong dan video"])
    wb.save(excel_path)


def log_closed_code(barcode, video_path):
    if not barcode or not video_path:
        return

    now = datetime.now()
    excel_path = get_excel_path(now)
    ensure_excel_file(excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active
    ws.append([now.strftime("%Y-%m-%d %H:%M:%S"), barcode, str(video_path)])
    wb.save(excel_path)


def is_barcode_closed(barcode):
    if not barcode:
        return False

    now = datetime.now()
    excel_path = get_excel_path(now)
    if not excel_path.exists():
        return False

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        val = row[0]
        if val is None:
            continue
        if str(val).strip() == str(barcode).strip():
            return True

    return False


def new_filename():
    global current_barcode

    now = datetime.now()
    ts = now.strftime("%Y%m%d_%H%M%S")
    video_dir = build_daily_dir(VIDEO_ROOT, now)
    video_dir.mkdir(parents=True, exist_ok=True)

    if current_barcode:
        name = f"{current_barcode}_{ts}.mp4"
        current_barcode = None
        return video_dir / name

    return video_dir / f"record_{ts}.mp4"

def start_ffmpeg():
    global ffmpeg_process, recording, current_video_path

    if recording:
        return

    filename = new_filename()

    cmd = [
        "ffmpeg", "-y",
        "-rtsp_transport", "tcp",
        "-i", rtsp_url,
        "-c", "copy",
        "-movflags", "+faststart",
        str(filename)
    ]

    ffmpeg_process = subprocess.Popen(
        cmd,
        stdin=subprocess.PIPE,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL
    )

    recording = True
    current_video_path = filename

    barcode_entry.config(state="normal")   # ✅ CHO BẮN
    barcode_entry.focus_set()

    status.config(text=f"🔴 Đang ghi: {filename.name}")

def start_record():
    global recording

    if recording:
        status.config(text="⚠️ Đang ghi rồi")
        return

    recording = True
    start_ffmpeg()

def cut_record():
    global ffmpeg_process, recording, current_video_path
    global pending_barcode, current_barcode
    # require a pending barcode and a running recording process
    if not pending_barcode:
        status.config(text="⚠️ Không có mã để đóng")
        return

    if not recording or not ffmpeg_process:
        status.config(text="⚠️ Chưa đang ghi")
        return

    code = str(pending_barcode).strip()
    if not code:
        pending_barcode = None
        status.config(text="⚠️ Mã rỗng")
        return

    # Must have a define character at the end
    define_char = code[-1]
    if define_char not in VALID_DEFINES:
        pending_barcode = None
        status.config(text="❌ Mã vạch không đúng dạng define")
        return

    # If this machine's define doesn't match, ignore (do not cut)
    if define_char != CURRENT_DEFINE:
        pending_barcode = None
        status.config(text=f"ℹ️ Mã define '{define_char}' không cho cắt trên máy này")
        return

    # OK: this machine should cut. Use the stripped code for logging/lookup
    stripped_code = code[:-1]

    closed_barcode = stripped_code
    closed_video_path = current_video_path

    current_barcode = closed_barcode
    pending_barcode = None

    try:
        ffmpeg_process.stdin.write(b"q")
        ffmpeg_process.stdin.flush()
    except:
        pass

    ffmpeg_process.wait()

    ffmpeg_process = None
    recording = False
    current_video_path = None

    log_closed_code(closed_barcode, closed_video_path)

    # resume recording immediately
    start_ffmpeg()



def stop_record():
    global ffmpeg_process, recording, current_video_path

    if not recording:
        return

    try:
        ffmpeg_process.stdin.write(b"q")
        ffmpeg_process.stdin.flush()
    except:
        pass

    ffmpeg_process.wait()

    ffmpeg_process = None
    recording = False
    current_video_path = None

    barcode_entry.config(state="disabled")  # ❌ KHÔNG CHO BẮN

    status.config(text="⏹️ Đã dừng ghi")

def on_barcode_enter(event):
    global pending_barcode

    if not recording:
        status.config(text="⚠️ Chưa bấm Bắt đầu ghi")
        barcode_entry.delete(0, tk.END)
        return

    code = barcode_entry.get().strip()
    if not code:
        return

    # Validate define char at end
    if code[-1] not in VALID_DEFINES:
        status.config(text="❌ Mã vạch không đúng dạng define")
        barcode_entry.delete(0, tk.END)
        return

    # Nếu mã đã được đóng trước đó -> thông báo và không xử lý
    if is_barcode_closed(code):
        status.config(text=f"⚠️ Mã vạch {code} đã đóng")
        barcode_entry.delete(0, tk.END)
        return

    pending_barcode = code
    barcode_entry.delete(0, tk.END)

    status.config(text=f"✅ Bắn thành công: {pending_barcode}")

    cut_record()


# ===== GUI =====
root = tk.Tk()
root.title("RTSP Recorder")

# ▶ Start
btn_start = tk.Button(root, text="▶ Bắt đầu ghi", command=start_ffmpeg)
btn_start.pack(pady=5)

# 📦 Input bắn mã vạch
barcode_entry = tk.Entry(root, font=("Arial", 14), state="disabled")
barcode_entry.pack(pady=8)
barcode_entry.bind("<Return>", on_barcode_enter)

# ⏹ Stop
btn_stop = tk.Button(root, text="⏹ Dừng hẳn", command=stop_record)
btn_stop.pack(pady=5)

status = tk.Label(root, text="⏸️ Chưa ghi")
status.pack(pady=10)

root.mainloop()


# ===== Background HTTP API =====
class ApiHandler(BaseHTTPRequestHandler):
    def _set_headers(self, code=200, content_type='application/json'):
        self.send_response(code)
        self.send_header('Content-Type', content_type)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_OPTIONS(self):
        self._set_headers()

    def do_POST(self):
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8') if length else ''
        try:
            data = json.loads(body) if body else {}
        except Exception:
            data = {}

        if self.path == '/fire':
            barcode = data.get('barcode')
            if not barcode:
                self._set_headers(400)
                self.wfile.write(json.dumps({'error': 'missing barcode'}).encode())
                return

            # validate define char on the posted barcode
            try:
                code = str(barcode).strip()
            except Exception:
                code = ''

            if not code or code[-1] not in VALID_DEFINES:
                self._set_headers(400)
                self.wfile.write(json.dumps({'error': 'Mã vạch không đúng dạng define'}).encode())
                return

            def do_fire():
                global pending_barcode
                pending_barcode = str(barcode)
                # ensure recording
                if not recording:
                    start_ffmpeg()
                    # wait a bit then cut
                    root.after(1200, cut_record)
                else:
                    cut_record()

            root.after(0, do_fire)
            self._set_headers(200)
            self.wfile.write(json.dumps({'status': 'ok', 'barcode': barcode}).encode())
            return

        if self.path == '/start':
            root.after(0, start_ffmpeg)
            self._set_headers(200)
            self.wfile.write(json.dumps({'status': 'started'}).encode())
            return

        if self.path == '/stop':
            root.after(0, stop_record)
            self._set_headers(200)
            self.wfile.write(json.dumps({'status': 'stopped'}).encode())
            return

        self._set_headers(404)
        self.wfile.write(json.dumps({'error': 'not found'}).encode())


def start_api_server(port=8765):
    # find free port if 8765 busy
    server = ThreadingHTTPServer(('127.0.0.1', port), ApiHandler)
    threading.Thread(target=server.serve_forever, daemon=True).start()
    return server


try:
    start_api_server()
except Exception:
    pass