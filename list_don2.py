import json
from datetime import datetime
from pathlib import Path

import webview
from openpyxl import load_workbook, Workbook
import threading
import urllib.parse
import re
import mimetypes
import os
import sys
import shlex
import time
import json
from datetime import datetime
from pathlib import Path

import webview
from openpyxl import load_workbook, Workbook
import threading
import urllib.parse
import re
import mimetypes
import os
import subprocess
import webbrowser
import shutil
from flask import Flask, request, Response, send_file, jsonify
from werkzeug.serving import make_server
from define_config import VALID_DEFINES

BASE_DIR = Path(__file__).resolve().parent
EXCEL_ROOT = BASE_DIR / "excel"
PAGE_SIZE = 20
VIDEO_ROOT = BASE_DIR / "video"
from rtsp_config import rtsp_url
ffmpeg_process = None
recording = False
current_video_path = None


def build_daily_dir(root_dir, dt):
    return root_dir / dt.strftime("%Y") / dt.strftime("%m") / dt.strftime("%d")


def new_filename():
    global current_video_path
    now = datetime.now()
    video_dir = build_daily_dir(VIDEO_ROOT, now)
    video_dir.mkdir(parents=True, exist_ok=True)
    ts = now.strftime("%Y%m%d_%H%M%S")
    filename = video_dir / f"record_{ts}.mp4"
    current_video_path = filename
    return filename


def get_today_excel_path():
    now = datetime.now()
    daily_dir = EXCEL_ROOT / now.strftime("%Y") / now.strftime("%m") / now.strftime("%d")
    return daily_dir / f"{now.strftime('%Y%m%d')}.xlsx"


def save_workbook_with_retry(workbook, path, retries=3, delay=0.5):
    """Try to save `workbook` to `path` with retries on PermissionError.

    On success returns the path (as a Path) where the workbook was saved.
    If the target is locked, attempts to save a copy with suffix `_copy_<ts>.xlsx`
    and returns that path. Raises other exceptions on failure.
    """
    last_exc = None
    for attempt in range(1, retries + 1):
        try:
            workbook.save(path)
            return Path(path)
        except PermissionError as e:
            last_exc = e
            time.sleep(delay)
        except Exception:
            # non-permission errors should surface immediately
            raise

    # retries exhausted — try saving to an alternate copy name in same folder
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    p = Path(path)
    copy_path = p.with_name(f"{p.stem}_copy_{ts}{p.suffix}")
    try:
        workbook.save(copy_path)
        return copy_path
    except Exception as e:
        raise PermissionError(f"Failed to save {path} after {retries} attempts: {last_exc}; copy failed: {e}")


def resolve_video_path(raw_path):
    video_path = Path(str(raw_path).strip())
    if not video_path.is_absolute():
        video_path = (BASE_DIR / video_path).resolve()
    return video_path


def parse_close_time(value):
    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return datetime.min

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return datetime.min


def load_today_orders():
    excel_path = get_today_excel_path()
    print(f"[DEBUG] load_today_orders: looking for excel at {excel_path}")
    if not excel_path.exists():
        print(f"[DEBUG] load_today_orders: excel not found: {excel_path}")
        return [], excel_path

    # load workbook for reading and possible updating user_count
    workbook = load_workbook(excel_path)
    worksheet = workbook.active

    rows = []
    # determine today's latest recorded video (if any) so we can show its name for all rows
    latest_today_video = None
    try:
        today_dir = build_daily_dir(VIDEO_ROOT, datetime.now())
        if today_dir.exists():
            vfiles = list(today_dir.glob('**/*.mp4'))
            if vfiles:
                vfiles.sort(key=lambda p: p.stat().st_mtime, reverse=True)
                latest_today_video = vfiles[0]
    except Exception:
        latest_today_video = None
    # Expect columns: close_time, barcode, user, user_count, video_path, elapsed_seconds
    for r_idx, row_cells in enumerate(worksheet.iter_rows(min_row=2, max_col=6), start=2):
        values = [c.value for c in row_cells]
        if not any(values):
            continue

        close_time = values[0] or ""
        barcode = str(values[1] or "")
        user_cell = values[2] or ""
        # compute user from barcode if cell empty
        user_char = str(user_cell) if user_cell else ''
        if not user_char and barcode:
            last = barcode.strip()[-1:]
            if last in VALID_DEFINES:
                user_char = last

        # stored_user_count = values[3] if len(values) > 3 else None
        video_raw = values[4] if len(values) > 4 and values[4] is not None else ""
        elapsed = values[5] if len(values) > 5 and values[5] is not None else ""
        sort_key = parse_close_time(close_time)

        resolved_video = resolve_video_path(video_raw)
        exists = resolved_video.exists()
        # Skip rows that do not reference an existing video file
        if not exists:
            continue

        rows.append({
            "close_time": str(close_time),
            "barcode": str(barcode),
            "user": user_char,
            "user_count": '',
            "video_path": str(resolved_video),
            # only set video_name when the referenced file actually exists
            "video_name": resolved_video.name,
            "elapsed": str(elapsed),
            "exists": exists,
            "sort_key": sort_key.isoformat(),
            "row_idx": r_idx,
        })

    # sort for display (newest first)
    rows.sort(key=lambda item: item["sort_key"], reverse=True)

    # recompute per-user sequential counts in display order and write back to excel
    counters = {}
    for item in rows:
        u = item.get('user') or ''
        if u:
            counters[u] = counters.get(u, 0) + 1
            item['user_count'] = str(counters[u])
        else:
            item['user_count'] = ''

    # compute start_time per user group: for each user, the largest `user_count` entry
    # is considered the first segment and starts at 0. Subsequent segments start at
    # the previous segment's end (previous elapsed). Store as string seconds.
    from collections import defaultdict
    groups = defaultdict(list)
    for item in rows:
        u = item.get('user') or ''
        if not u:
            continue
        # parse user_count and elapsed if possible
        try:
            uc = int(item.get('user_count')) if (item.get('user_count') not in (None, '')) else None
        except Exception:
            uc = None
        try:
            elapsed_f = float(item.get('elapsed')) if (item.get('elapsed') not in (None, '')) else None
        except Exception:
            elapsed_f = None
        groups[u].append({'item': item, 'user_count': uc, 'elapsed': elapsed_f})

    for u, lst in groups.items():
        # sort by user_count desc (largest STT first), fallback to elapsed desc
        lst.sort(key=lambda x: (-(int(x['user_count']) if x['user_count'] is not None else 0), -(float(x['elapsed']) if x['elapsed'] is not None else 0)))
        prev_end = 0.0
        for rec in lst:
            it = rec['item']
            if rec['elapsed'] is None:
                it['start_time'] = ''
                continue
            it['start_time'] = str(prev_end)
            prev_end = float(rec['elapsed'])

    # write back user_count to worksheet (column 4)
    try:
        for item in rows:
            ws_row = item.get('row_idx')
            if ws_row:
                workbook.active.cell(row=ws_row, column=4, value=item.get('user_count') or '')
        workbook.save(excel_path)
    except Exception:
        pass

    for item in rows:
        item.pop("sort_key", None)
        # keep row_idx so client can reference the Excel row for updates

    print(f"[DEBUG] load_today_orders: loaded {len(rows)} rows from {excel_path}")
    if len(rows) > 0:
        sample = rows[:5]
        try:
            print(f"[DEBUG] sample rows: {json.dumps(sample, default=str) }")
        except Exception:
            pass
    return rows, excel_path


def append_order(barcode_value, elapsed_seconds=None):
    excel_path = get_today_excel_path()
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if not excel_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["close_time", "barcode", "user", "user_count", "video_path", "elapsed_seconds"])
        try:
            saved_path = save_workbook_with_retry(wb, excel_path, retries=5, delay=0.5)
            if str(saved_path) != str(excel_path):
                raise Exception(f'failed saving excel to target (file locked) — saved copy: {saved_path}')
        except PermissionError as pe:
            raise Exception(f'failed saving excel (permission denied): {pe}')

    wb = load_workbook(excel_path)
    ws = wb.active
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    el = '' if elapsed_seconds is None else float(elapsed_seconds)
    b = str(barcode_value or "")
    user_char = ''
    if b:
        last = b.strip()[-1:]
        if last in VALID_DEFINES:
            user_char = last

    # compute next user_count by scanning existing rows
    next_count = ''
    if user_char:
        existing = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
        cnt = sum(1 for v in existing if v == user_char)
        next_count = str(cnt + 1)

    ws.append([now, b, user_char, next_count, "", el])
    wb.save(excel_path)


def build_html(rows, excel_path, base_url):
    # Use a template with placeholders to avoid f-string brace interpolation issues
    # dump JSON and escape closing script tag to avoid breaking the surrounding <script>
    data_json = json.dumps(rows, ensure_ascii=False)
    data_json = data_json.replace('</script>', '<\\/script>')
    template = """
<!doctype html>
<html>
<head>
  <meta charset="utf-8" />
  <title>Danh sach don hom nay</title>
    <style>
    :root {--bg: #f4efe4;--surface: #fff8ee;--ink: #1f2328;--muted: #5f676f;--accent: #005f73;--accent-2: #ee9b00;--line: #d9cbb0;}
    * { box-sizing: border-box; }
    body { margin: 0; font-family: "Segoe UI", Tahoma, sans-serif; color: var(--ink); background: radial-gradient(circle at 15% 10%, #fff5dd 0%, var(--bg) 45%); min-height: 100vh; }
    .wrap { max-width: 1400px; margin: 24px auto; padding: 0 16px 24px; }
    .head { background: var(--surface); border: 1px solid var(--line); border-radius: 14px; padding: 14px 16px; margin-bottom: 14px; }
    .title { margin: 0; font-size: 24px; color: var(--accent); }
    .meta { margin-top: 6px; color: var(--muted); font-size: 13px; word-break: break-all; }
    .grid { display: block; gap: 14px; }
    @media (min-width: 980px) { .grid { grid-template-columns: 1fr; } }
    .card { background: var(--surface); border: 1px solid var(--line); border-radius: 14px; padding: 12px; box-shadow: 0 8px 20px rgba(31, 35, 40, 0.06); }
    table { width: 100%; border-collapse: collapse; font-size: 14px; table-layout: fixed; }
    th, td { border-bottom: 1px solid var(--line); text-align: left; padding: 10px 8px; vertical-align: top; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
    /* limit video name width and narrow the final Video column */
    table th:nth-child(7), table td:nth-child(7) { width: 260px; max-width: 260px; }
    table th:nth-child(8), table td:nth-child(8) { width: 90px; max-width: 90px; text-align: center; }
    tr.selected { background: rgba(0,95,115,0.06); }
    th { color: var(--accent); font-weight: 700; background: #fff2dc; position: sticky; top: 0; }
    .table-wrap { max-height: 80vh; overflow: auto; border: 1px solid var(--line); border-radius: 10px; background: #fff; }
    .btn { border: 0; border-radius: 8px; padding: 7px 10px; background: var(--accent); color: #fff; cursor: pointer; font-weight: 600; }
    .btn:disabled { background: #8a8f95; cursor: not-allowed; }
    .pager { margin-top: 10px; display: flex; align-items: center; gap: 8px; }
    .pager-info { color: var(--muted); font-size: 13px; margin-left: 4px; }
    .video-box { border: 1px dashed var(--line); border-radius: 10px; padding: 10px; min-height: 240px; background: #fff; }
    .video-title { margin: 0 0 8px; color: var(--accent); font-size: 15px; }
    #videoInfo { font-size: 12px; color: var(--muted); word-break: break-all; margin: 0 0 8px; }
    video { width: 100%; max-height: 420px; background: #000; border-radius: 8px; }
    .empty { color: var(--muted); font-style: italic; }
    .scan-row { display:flex; gap:8px; margin-bottom:8px; }
    .scan-input { flex:1; padding:8px; border-radius:8px; border:1px solid var(--line); font-size:14px; }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="head">
      <h1 class="title">Danh sach don hom nay</h1>
      <div class="meta">Excel: %%EXCEL_PATH%%</div>
      <div class="meta" id="count"></div>
    </div>

    <div class="grid">
      <section class="card">
                <div class="scan-row">
                    <input id="barcodeInput" class="scan-input" placeholder="Bắn mã vạch ở đây rồi Enter..." />
                    <button class="btn" id="startBtn">Bắt đầu</button>
                                        <button class="btn" id="stopBtn" disabled> Dừng </button>
                                        <button class="btn" id="endBtn"> Kết thúc </button>
                                        <button class="btn" id="finalizeBtn"> Hoàn tất quy trình </button>
                    <span id="timerDisplay" style="align-self:center;margin-left:8px;color:var(--muted)">0.0s</span>
                    <button class="btn" id="scanBtn">Bắn</button>
                </div>
                <div class="table-wrap">
          <table>
            <thead>
                                                                <tr>
                                                                    <th>Thoi gian dong</th>
                                                                    <th>Ma vach</th>
                                                                    <th>Người dùng</th>
                                                                    <th>STT</th>
                                                                    <th>Thời gian bắt đầu</th>
                                                                    <th>Thời gian đếm</th>
                                                                    <th>Tên video</th>
                                                                    <th>Video</th>
                                                                </tr>
            </thead>
            <tbody id="rows"></tbody>
          </table>
                </div>
                <div id="rowActions" style="margin:10px 0"></div>
                <div class="pager">
                    <button class="btn" id="prevBtn" onclick="prevPage()">Trang truoc</button>
                    <button class="btn" id="nextBtn" onclick="nextPage()">Trang sau</button>
                    <span class="pager-info" id="pageInfo">Trang 1/1</span>
                </div>
                <!-- Video viewer moved below the table -->
                <div class="video-box" style="margin-top:12px">
                        <h3 class="video-title">Xem video don</h3>
                        <p id="videoInfo">Chon mot don de xem video.</p>
                        <video id="player" controls></video>
                        <div style="margin-top:10px">
                                <h4 style="margin:6px 0 8px;color:var(--accent)">Danh sách Excel hôm nay</h4>
                                <div id="excelList" style="font-size:13px;color:var(--muted)">Đang tải...</div>
                        </div>
                </div>
            </section>
    </div>
  </div>

  <script>
    let data = %%DATA_JSON%%;
        console.log('Loaded data from server (embedded):', (data && data.length) ? data.length : 0, data && data.slice ? data.slice(0,5) : data);
    const pageSize = %%PAGE_SIZE%%;
    let currentPage = 1;
    let selectedIndex = null;
    const rowsEl = document.getElementById('rows');
    const countEl = document.getElementById('count');
    const player = document.getElementById('player');
    const info = document.getElementById('videoInfo');
    const rowActionsEl = document.getElementById('rowActions');
    const pageInfoEl = document.getElementById('pageInfo');
    const prevBtn = document.getElementById('prevBtn');
    const nextBtn = document.getElementById('nextBtn');
    const scanInput = document.getElementById('barcodeInput');
    const scanBtn = document.getElementById('scanBtn');
    const startBtn = document.getElementById('startBtn');
    const stopBtn = document.getElementById('stopBtn');
    const endBtn = document.getElementById('endBtn');
    const timerDisplay = document.getElementById('timerDisplay');

    let timerStart = null;
    let timerInterval = null;

    // disable scanning until user presses Start
    scanInput.disabled = true;
    scanBtn.disabled = true;
    // disable End until Start is pressed
    endBtn.disabled = true;

    function formatSeconds(s) {
      if (s === null || s === undefined || isNaN(s)) return '';
      return (Math.round(s * 1000) / 1000) + 's';
    }

    function startTimer() {
      if (timerInterval) clearInterval(timerInterval);
      timerStart = Date.now();
      startBtn.textContent = 'Đang...';
      startBtn.disabled = true;
      stopBtn.disabled = false;
            // allow End after starting
            endBtn.disabled = false;
      timerInterval = setInterval(() => {
        const elapsed = (Date.now() - timerStart) / 1000;
        timerDisplay.textContent = formatSeconds(elapsed);
      }, 200);
        // enable scanning input when started
        scanInput.disabled = false;
        scanBtn.disabled = false;
        scanInput.focus();
    }

    function stopTimer() {
      if (timerInterval) { clearInterval(timerInterval); timerInterval = null; }
      // freeze display; allow starting a fresh timer
      startBtn.textContent = 'Bắt đầu';
      startBtn.disabled = false;
      stopBtn.disabled = true;
      timerStart = null;
        // disable scanning after stop
        scanInput.disabled = true;
        scanBtn.disabled = true;
    }

    function playVideo(item) {
      var url = item.video_uri || item.video_path || '';
      if (!url) {
        info.textContent = 'Khong tim thay file video.';
        return;
      }
      if (window.pywebview && window.pywebview.api && window.pywebview.api.open_in_browser) {
        window.pywebview.api.open_in_browser(url);
      } else {
        window.open(url, '_blank');
      }
      info.textContent = item.video_path;
    }

        function selectRow(i) {
            selectedIndex = i;
            renderTable();
            renderRowActions();
        }

        function renderRowActions() {
            if (!rowActionsEl) return;
            if (selectedIndex === null || typeof data[selectedIndex] === 'undefined') {
                rowActionsEl.innerHTML = '';
                return;
            }
            const item = data[selectedIndex];
            const disabled = item.exists ? '' : 'disabled';
            const btn = '<button class="btn" ' + disabled + ' onclick="playVideo(data[' + selectedIndex + '])">Xem</button>';
            rowActionsEl.innerHTML = '<div style="display:flex;align-items:center;gap:8px">' + btn + '<div style="color:var(--muted)">' + (item.video_name || item.video_path || '') + '</div></div>';
        }

    function renderTable() {
                        if (data.length === 0) {
                                                                rowsEl.innerHTML = '<tr><td colspan="8" class="empty">Khong co du lieu don cho ngay hom nay.</td></tr>';
        countEl.textContent = 'Tong don: 0';
        pageInfoEl.textContent = 'Trang 0/0';
        prevBtn.disabled = true;
        nextBtn.disabled = true;
        return;
      }
      const totalPages = Math.ceil(data.length / pageSize);
      if (currentPage > totalPages) currentPage = totalPages;
      const start = (currentPage - 1) * pageSize;
      const end = start + pageSize;
      const pageRows = data.slice(start, end);
      countEl.textContent = 'Tong don: ' + data.length;
      pageInfoEl.textContent = 'Trang ' + currentPage + '/' + totalPages;
      prevBtn.disabled = currentPage <= 1;
      nextBtn.disabled = currentPage >= totalPages;
                        rowsEl.innerHTML = pageRows.map((item, index) => {
                const actualIndex = start + index;
                const disabled = item.exists ? '' : 'disabled';
                const selectedClass = (actualIndex === selectedIndex) ? 'selected' : '';
                    const cutBtn = '<button class="btn" ' + disabled + ' onclick="cutManual(' + actualIndex + ')">Cắt</button>';
                    const viewBtn = '<button class="btn" ' + disabled + ' onclick="playVideo(data[' + actualIndex + '])">Xem</button>';
                                                                    return '<tr class="' + selectedClass + '" onclick="selectRow(' + actualIndex + ')">' + '<td>' + item.close_time + '</td>' + '<td>' + item.barcode + '</td>' + '<td>' + (item.user || '') + '</td>' + '<td>' + (item.user_count || '') + '</td>' + '<td>' + (item.start_time || '') + '</td>' + '<td>' + (item.elapsed || '') + '</td>' + '<td>' + (item.video_name || '') + '</td>' + '<td>' + cutBtn + ' ' + viewBtn + '</td>' + '</tr>';
            }).join('');
                        renderRowActions();
    }

    function nextPage() { const totalPages = Math.ceil(data.length / pageSize); if (currentPage < totalPages) { currentPage += 1; renderTable(); } }
    function prevPage() { if (currentPage > 1) { currentPage -= 1; renderTable(); } }

    async function refreshData() {
      try { const resp = await fetch('/data'); if (!resp.ok) return; const json = await resp.json(); data = json; currentPage = 1; renderTable(); } catch (e) { console.error('Refresh error', e); }
    }

    async function sendBarcode(barcode) {
            try {
                if (!timerStart) {
                    alert('Vui lòng nhấn Bắt đầu trước khi bắn mã vạch.');
                    return;
                }
                let elapsed = null;
                elapsed = (Date.now() - timerStart) / 1000;
                const resp = await fetch('/add', { method: 'POST', headers: { 'Content-Type': 'application/json' }, body: JSON.stringify({ barcode, elapsed_seconds: elapsed }) });
                if (resp.ok) {
                    scanInput.value = ''; scanInput.focus();
                    // do NOT stop or reset the timer here; timer keeps running until user presses Stop
                    await refreshData();
                } else { alert('Lỗi khi gửi mã vạch'); }
            } catch (e) { alert('Lỗi kết nối: ' + e.message); }
    }

    scanBtn.addEventListener('click', () => { const v = scanInput.value.trim(); if (!v) return; sendBarcode(v); });
    scanInput.addEventListener('keydown', (e) => { if (e.key === 'Enter') { const v = scanInput.value.trim(); if (!v) return; sendBarcode(v); } });
    startBtn.addEventListener('click', async () => {
      startTimer();
      try {
        await fetch('/rec_start', { method: 'POST' });
      } catch (e) {
        console.error('Start recording proxy failed', e);
      }
    });

    stopBtn.addEventListener('click', async () => {
      stopTimer();
      try {
                const resp = await fetch('/rec_stop', { method: 'POST' });
                if (resp.ok) {
                    await refreshData();
                    const j = await resp.json();
                    if (j && j.video) console.info('Ghi video xong: ' + j.video);
                }
      } catch (e) {
        console.error('Stop recording proxy failed', e);
      }
    });
                endBtn.addEventListener('click', async () => {
                        if (!confirm('Bạn có chắc muốn cắt các đoạn video theo các stt người dùng hiện có không?')) return;
                        endBtn.disabled = true;
                        try {
                                const resp = await fetch('/cut_end', { method: 'POST' });
                                if (!resp.ok) {
                                        const txt = await resp.text();
                                        alert('Lỗi khi cắt video: ' + txt);
                                } else {
                                        const j = await resp.json();
                                        alert('Hoàn tất cắt ' + (j.files ? j.files.length : 0) + ' file.');
                                        await refreshData();
                                }
                        } catch (e) {
                                alert('Lỗi khi kết nối: ' + e.message);
                        } finally {
                                endBtn.disabled = false;
                        }
                });

        finalizeBtn.addEventListener('click', async () => {
            if (!confirm('Bạn có chắc muốn hoàn tất quy trình và lưu file Excel phiên này không?')) return;
            finalizeBtn.disabled = true;
            try {
                const resp = await fetch('/finalize', { method: 'POST' });
                if (!resp.ok) {
                    const txt = await resp.text();
                    alert('Lỗi khi hoàn tất: ' + txt);
                } else {
                    const j = await resp.json();
                    alert('Hoàn tất lưu: ' + (j.file ? j.file : 'Không có file mới'));
                    await refreshExcelList();
                }
            } catch (e) {
                alert('Lỗi khi kết nối: ' + e.message);
            } finally {
                finalizeBtn.disabled = false;
            }
        });

        async function refreshExcelList() {
            try {
                const resp = await fetch('/excel_list');
                if (!resp.ok) { document.getElementById('excelList').textContent = 'Không thể tải danh sách.'; return; }
                const list = await resp.json();
                const el = document.getElementById('excelList');
                if (!list || list.length === 0) { el.innerHTML = '<div class="empty">Không có file Excel hôm nay.</div>'; return; }
                el.innerHTML = '<ul style="margin:0;padding-left:18px">' + list.map(i => '<li><a href="' + i.url + '" target="_blank">' + i.name + '</a></li>').join('') + '</ul>';
            } catch (e) {
                document.getElementById('excelList').textContent = 'Lỗi tải: ' + e.message;
            }
        }

        async function cutManual(index) {
            const item = data[index];
            if (!item) { alert('Không tìm thấy mục'); return; }
            const payload = { row: item.row_idx, video_path: item.video_path, start: item.start_time || 0, end: item.elapsed || 0, barcode: item.barcode };
            try {
                // request a preview of the cut command from server
                let cmdText = '';
                try {
                    const pre = await fetch('/preview_cut', { method: 'POST', headers: { 'Content-Type': 'application/json' }, body: JSON.stringify(payload) });
                    if (pre.ok) {
                        const pj = await pre.json();
                        cmdText = pj.cmd || '';
                    } else {
                        const txt = await pre.text();
                        cmdText = 'Preview error: ' + txt;
                    }
                } catch (err) {
                    cmdText = 'Preview failed: ' + (err.message || err);
                }

                if (!confirm(`Cắt đoạn cho mã: ${item.barcode || ''}?\n\nLệnh sẽ chạy:\n${cmdText}`)) return;

                const resp = await fetch('/cut_manual', { method: 'POST', headers: { 'Content-Type': 'application/json' }, body: JSON.stringify(payload) });
                const j = await resp.json();
                if (!resp.ok) {
                    alert('Lỗi cắt: ' + (j.error || 'Unknown'));
                } else {
                    let msg = 'Cắt xong: ' + (j.file || '');
                    if (j.excel_copy) msg += '\\nExcel saved as copy: ' + j.excel_copy;
                    alert(msg);
                    await refreshData();
                }
            } catch (e) { alert('Lỗi kết nối: ' + e.message); }
        }

        // initial load of excel list
        refreshExcelList();

        renderTable();
  </script>
</body>
</html>
"""

    html = template.replace('%%DATA_JSON%%', data_json).replace('%%PAGE_SIZE%%', str(PAGE_SIZE)).replace('%%EXCEL_PATH%%', str(excel_path))
    return html


def main():
    VIDEO_ROOT.mkdir(parents=True, exist_ok=True)

    app = Flask(__name__)

    class Api:
        def open_in_browser(self, url):
            try:
                if not url:
                    return False
                chrome_paths = [
                    os.path.join(os.environ.get('PROGRAMFILES', 'C:\\Program Files'), 'Google\\Chrome\\Application\\chrome.exe'),
                    os.path.join(os.environ.get('PROGRAMFILES(X86)', 'C:\\Program Files (x86)'), 'Google\\Chrome\\Application\\chrome.exe'),
                    os.path.join(os.environ.get('LOCALAPPDATA', os.path.expanduser('~\\AppData\\Local')), 'Google\\Chrome\\Application\\chrome.exe'),
                ]
                for ch in chrome_paths:
                    if os.path.exists(ch):
                        try:
                            subprocess.Popen([ch, f'--app={url}', '--window-size=900,600'], shell=False)
                            return True
                        except Exception:
                            try:
                                subprocess.Popen([ch, '--new-window', url, '--window-size=900,600'], shell=False)
                                return True
                            except Exception:
                                continue

                webbrowser.open(url)
                return True
            except Exception:
                try:
                    webbrowser.open(url)
                    return True
                except Exception:
                    return False


    @app.route('/')
    def index():
        # load rows fresh each request so newly appended orders appear after reload
        rows, excel_path = load_today_orders()

        # convert video paths to flask URLs served by /video/<path:filename>
        for item in rows:
            if item.get('exists'):
                try:
                    p = Path(item['video_path'])
                    rel = p.relative_to(VIDEO_ROOT)
                    quoted = "/".join(urllib.parse.quote(part) for part in rel.parts)
                    item['video_uri'] = base_url + f"/video/{quoted}"
                except Exception:
                    try:
                        item['video_uri'] = Path(item['video_path']).as_uri()
                    except Exception:
                        item['video_uri'] = ""
            else:
                item['video_uri'] = ""

        html = build_html(rows, excel_path, base_url)
        return html


    def start_ffmpeg():
        global ffmpeg_process, recording, current_video_path

        if recording:
            return

        filename = new_filename()

        cmd = [
            "ffmpeg", "-y",
            "-rtsp_transport", "tcp",
            "-i", rtsp_url,
            "-c", "copy",
            "-movflags", "+faststart",
            str(filename)
        ]

        ffmpeg_process = subprocess.Popen(
            cmd,
            stdin=subprocess.PIPE,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )

        recording = True

    def start_record():
        global recording

        if recording:
            return None

        start_ffmpeg()
        return current_video_path

    def stop_record():
        global ffmpeg_process, recording, current_video_path

        if not recording:
            return

        try:
            ffmpeg_process.stdin.write(b"q")
            ffmpeg_process.stdin.flush()
        except:
            pass

        ffmpeg_process.wait()

        ffmpeg_process = None
        recording = False
        # do not clear current_video_path here; let caller (rec_stop) attach it to Excel
        return
    


    @app.route('/add', methods=['POST'])
    def add():
        try:
            data = request.get_json(force=True)
            barcode = data.get('barcode') if isinstance(data, dict) else None
            elapsed = data.get('elapsed_seconds') if isinstance(data, dict) else None
            if not barcode:
                return jsonify({'ok': False, 'error': 'missing barcode'}), 400
            append_order(barcode, elapsed)
            return jsonify({'ok': True})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500


    def find_latest_video_file():
        now = datetime.now()
        video_dir = VIDEO_ROOT / now.strftime("%Y") / now.strftime("%m") / now.strftime("%d")
        if not video_dir.exists():
            return None
        files = list(video_dir.glob('**/*.mp4'))
        if not files:
            return None
        files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return files[0]


    def find_video_from_excel():
        """Return the most recent existing video path recorded in today's Excel (column 5), or None."""
        excel_path = get_today_excel_path()
        if not excel_path.exists():
            return None
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            # iterate bottom->top to prefer latest entries
            for r in range(ws.max_row, 1, -1):
                val = ws.cell(row=r, column=5).value
                if not val:
                    continue
                p = Path(str(val))
                if not p.is_absolute():
                    p = (BASE_DIR / p).resolve()
                if p.exists():
                    return p
        except Exception:
            return None
        return None


    def attach_latest_video_to_last_row(video_path: Path):
        excel_path = get_today_excel_path()
        if not excel_path.exists():
            return False
        wb = load_workbook(excel_path)
        ws = wb.active
        # write the given video_path into column 5 for all data rows (row 2..max_row)
        updated = 0
        for r in range(2, ws.max_row + 1):
            try:
                ws.cell(row=r, column=5, value=str(video_path))
                updated += 1
            except Exception:
                continue

        try:
            # try to save using the robust saver (handles locked Excel by creating a copy)
            saved_path = save_workbook_with_retry(wb, excel_path, retries=5, delay=0.5)
            # if saved to a different path, log and still return the count
        except PermissionError:
            # failed to save after retries/copy
            return 0

        return updated


    @app.route('/rec_start', methods=['POST'])
    def rec_start():
        try:
            filename = start_record()
            return jsonify({'ok': True, 'video': str(filename) if filename else None})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500


    @app.route('/rec_stop', methods=['POST'])
    def rec_stop():
        try:
            stop_record()
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500

        # after stopping, find latest video and attach
        latest = find_latest_video_file()
        if latest:
            attached_count = attach_latest_video_to_last_row(latest)
            try:
                # clear the in-memory current_video_path after successful attach
                global current_video_path
                current_video_path = None
            except Exception:
                pass
            return jsonify({'ok': True, 'video': str(latest), 'attached_count': attached_count})
        return jsonify({'ok': True, 'video': None, 'attached_count': 0})


    @app.route('/cut_end', methods=['POST'])
    def cut_end():
        try:
            # prefer explicit video path stored in Excel (column 5);
            # fallback to latest file in today's folder
            latest = find_video_from_excel() or find_latest_video_file()
            if not latest:
                return jsonify({'ok': False, 'error': 'No video found for today'}), 404

            files = cut_segments_from_video(latest)
            return jsonify({'ok': True, 'files': [str(p) for p in files]})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500


    @app.route('/cut_manual', methods=['POST'])
    def cut_manual():
        try:
            data = request.get_json(force=True)
            # expect: row (excel row index), video_path, start, end, barcode
            row = data.get('row')
            video_path = data.get('video_path')
            start = data.get('start')
            end = data.get('end')
            barcode = data.get('barcode')

            if not video_path or start is None or end is None or row is None:
                return jsonify({'ok': False, 'error': 'missing parameters'}), 400

            try:
                start_f = float(start)
                end_f = float(end)
            except Exception:
                return jsonify({'ok': False, 'error': 'invalid start/end'}), 400

            inp = Path(str(video_path))
            if not inp.is_absolute():
                inp = (BASE_DIR / inp).resolve()
            if not inp.exists():
                return jsonify({'ok': False, 'error': 'video file not found'}), 404

            # build output name: barcode without last char + _timestamp.mp4
            code = str(barcode or '')
            base_name = code[:-1] if len(code) > 0 else 'cut'
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            out_name = f"{base_name}_{ts}.mp4"
            out_dir = inp.parent
            out_path = out_dir / out_name

            # call cutvideo.py
            cutter = Path(__file__).resolve().parent / 'cutvideo.py'
            cmd = [sys.executable, str(cutter), str(inp), str(start_f), str(end_f), str(out_path)]
            # log the command for inspection
            try:
                print('Running cut command:', ' '.join(map(str, cmd)))
            except Exception:
                pass
            try:
                subprocess.run(cmd, check=True)
            except subprocess.CalledProcessError as e:
                return jsonify({'ok': False, 'error': f'cut failed: {e}'}), 500

            # write output path to Excel under header 'cutvideo' (create column if needed)
            excel_path = get_today_excel_path()
            if not excel_path.exists():
                return jsonify({'ok': False, 'error': 'excel not found'}), 404

            wb = load_workbook(excel_path)
            ws = wb.active
            # find or create 'cutvideo' header
            cut_col = None
            for c in range(1, ws.max_column + 1):
                h = ws.cell(row=1, column=c).value
                if isinstance(h, str) and h.strip().lower() == 'cutvideo':
                    cut_col = c
                    break
            if cut_col is None:
                cut_col = ws.max_column + 1
                ws.cell(row=1, column=cut_col, value='cutvideo')

            try:
                target_row = int(row)
                ws.cell(row=target_row, column=cut_col, value=str(out_path))
                try:
                    saved_path = save_workbook_with_retry(wb, excel_path, retries=5, delay=0.5)
                except PermissionError as pe:
                    return jsonify({'ok': False, 'error': f'failed writing excel (permission denied): {pe}'}), 500
                # if saved to a different path, include it in response
                if str(saved_path) != str(excel_path):
                    return jsonify({'ok': True, 'file': str(out_path), 'excel_copy': str(saved_path)})
            except Exception as e:
                return jsonify({'ok': False, 'error': 'failed writing excel: ' + str(e)}), 500

            return jsonify({'ok': True, 'file': str(out_path)})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500


    @app.route('/preview_cut', methods=['POST'])
    def preview_cut():
        try:
            data = request.get_json(force=True)
            row = data.get('row')
            video_path = data.get('video_path')
            start = data.get('start')
            end = data.get('end')
            barcode = data.get('barcode')

            if not video_path or start is None or end is None or row is None:
                return jsonify({'ok': False, 'error': 'missing parameters'}), 400

            try:
                start_f = float(start)
                end_f = float(end)
            except Exception:
                return jsonify({'ok': False, 'error': 'invalid start/end'}), 400

            inp = Path(str(video_path))
            if not inp.is_absolute():
                inp = (BASE_DIR / inp).resolve()
            if not inp.exists():
                return jsonify({'ok': False, 'error': 'video file not found'}), 404

            # build output name as in cut_manual
            code = str(barcode or '')
            base_name = code[:-1] if len(code) > 0 else 'cut'
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            out_name = f"{base_name}_{ts}.mp4"
            out_dir = inp.parent
            out_path = out_dir / out_name

            cutter = Path(__file__).resolve().parent / 'cutvideo.py'
            cmd = [sys.executable, str(cutter), str(inp), str(start_f), str(end_f), str(out_path)]
            # return a readable command string for confirmation
            cmd_str = ' '.join([shlex.quote(str(p)) for p in cmd]) if 'shlex' in globals() else ' '.join(map(str, cmd))
            return jsonify({'ok': True, 'cmd': cmd_str})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500


    def cut_segments_from_video(video_path: Path):
        """Cut the given full recording into per-user sequential segments and write paths back to Excel.

        For each user, cuts are created in ascending `user_count` order. The first segment for a user
        starts at 0.0 and ends at the user's first `elapsed_seconds`. Subsequent segments go from the
        previous elapsed to the current elapsed. The generated files are saved to a `cuts` folder in
        the same day directory as the original video and the path is written to column 5 of the Excel.
        Returns a list of output Path objects.
        """
        excel_path = get_today_excel_path()
        if not excel_path.exists():
            raise Exception('Excel file for today not found')

        wb = load_workbook(excel_path)
        ws = wb.active

        # collect rows that have a user and an elapsed_seconds value
        rows = []
        for r in range(2, ws.max_row + 1):
            user = ws.cell(row=r, column=3).value
            elapsed = ws.cell(row=r, column=6).value
            if not user or elapsed is None:
                continue
            try:
                elapsed_f = float(elapsed)
            except Exception:
                continue
            try:
                user_count_val = ws.cell(row=r, column=4).value
                user_count_int = int(user_count_val) if user_count_val not in (None, '') else None
            except Exception:
                user_count_int = None

            rows.append({'row': r, 'user': str(user), 'user_count': user_count_int, 'elapsed': elapsed_f})

        if not rows:
            return []

        from collections import defaultdict
        groups = defaultdict(list)
        for it in rows:
            groups[it['user']].append(it)

        output_files = []
        nowstamp = datetime.now().strftime('%y%m%d%H%M%S')

        for user, items in groups.items():
            # sort by user_count descending (largest STT first). If user_count
            # missing, treat as 0 and fall back to elapsed descending.
            def sort_key(x):
                uc = x.get('user_count')
                try:
                    ucv = int(uc) if uc is not None else 0
                except Exception:
                    ucv = 0
                return (-ucv, -float(x['elapsed']))

            items.sort(key=sort_key)

            # The first (largest stt) cut starts at 0. Subsequent cuts start
            # at the end time of the previous cut (prev_end) and end at the
            # current row's elapsed. This creates non-overlapping sequential
            # segments per the user's request.
            prev_end = 0.0
            for it in items:
                start = float(prev_end)
                end = float(it['elapsed'])
                duration = end - start
                if duration <= 0:
                    # advance prev_end to this elapsed even if no segment
                    prev_end = end
                    continue

                out_dir = video_path.parent / 'cuts'
                out_dir.mkdir(parents=True, exist_ok=True)
                safe_user = re.sub(r'[^A-Za-z0-9_-]', '_', it['user'])
                uc = it['user_count'] if it['user_count'] is not None else 0
                out_name = f"cut_{safe_user}_{uc}_{nowstamp}.mp4"
                out_path = out_dir / out_name

                # ffmpeg: seek to start, copy duration
                cmd = [
                    'ffmpeg', '-y', '-ss', str(start), '-i', str(video_path), '-t', str(duration), '-c', 'copy', str(out_path)
                ]
                try:
                    subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=True)
                except Exception:
                    # if direct copy fails, try a re-encode as a fallback
                    cmd2 = [
                        'ffmpeg', '-y', '-ss', str(start), '-i', str(video_path), '-t', str(duration), '-c:v', 'libx264', '-c:a', 'aac', str(out_path)
                    ]
                    subprocess.run(cmd2, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

                # write the path back to the excel (column 5)
                ws.cell(row=it['row'], column=5, value=str(out_path))
                output_files.append(out_path)
                prev = end

        wb.save(excel_path)
        return output_files


    @app.route('/finalize', methods=['POST'])
    def finalize():
        try:
            src = get_today_excel_path()
            if not src.exists():
                return jsonify({'ok': False, 'error': 'Excel file for today not found'}), 404

            # load workbook to validate that no close_time cells are empty
            wb = load_workbook(src)
            ws = wb.active
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=1).value
                if v is None or str(v).strip() == '':
                    return jsonify({'ok': False, 'error': 'Some close_time cells are empty; not saving copy'}), 400

            d = src.parent
            base = src.stem  # YYYYMMDD
            # find existing numbered copies
            nums = []
            for p in d.glob(f"{base}*.xlsx"):
                m = re.match(rf"{re.escape(base)}(?:_(\d+))?\.xlsx$", p.name)
                if m:
                    nums.append(int(m.group(1)) if m.group(1) else 0)
            next_num = 1
            if nums:
                next_num = max(nums) + 1
            dest = d / f"{base}_{next_num}.xlsx"

            # create a copy with the close_time column cleared (so original remains intact)
            # we already have the workbook loaded from src; clear column 1 for data rows
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=1, value='')

            wb.save(dest)
            files = [p.name for p in sorted(d.glob(f"{base}*.xlsx"))]
            return jsonify({'ok': True, 'file': dest.name, 'files': files})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500


    @app.route('/excel_list')
    def excel_list():
        excel_path = get_today_excel_path()
        d = excel_path.parent
        if not d.exists():
            return jsonify([])
        files = [p.name for p in sorted(d.glob(f"{excel_path.stem}*.xlsx"))]
        items = []
        for name in files:
            url = base_url + f"/excel/{urllib.parse.quote(name)}"
            items.append({'name': name, 'url': url})
        return jsonify(items)


    @app.route('/excel/<path:filename>')
    def excel(filename):
        safe_path = (EXCEL_ROOT / filename).resolve()
        try:
            safe_path.relative_to(EXCEL_ROOT.resolve())
        except Exception:
            return Response('Forbidden', status=403)
        if not safe_path.exists():
            return Response('File not found', status=404)
        return send_file(str(safe_path))


    @app.route('/data')
    def data_api():
        rows, _ = load_today_orders()
        print(f"[DEBUG] /data returning {len(rows)} rows")
        for item in rows:
            if item.get('exists'):
                try:
                    p = Path(item['video_path'])
                    rel = p.relative_to(VIDEO_ROOT)
                    quoted = "/".join(urllib.parse.quote(part) for part in rel.parts)
                    item['video_uri'] = base_url + f"/video/{quoted}"
                except Exception:
                    try:
                        item['video_uri'] = Path(item['video_path']).as_uri()
                    except Exception:
                        item['video_uri'] = ""
            else:
                item['video_uri'] = ""
        return jsonify(rows)


    def send_file_partial(path):
        file_path = Path(path)
        if not file_path.exists():
            return Response('File not found', status=404)

        file_size = file_path.stat().st_size
        range_header = request.headers.get('Range', None)
        if not range_header:
            return send_file(str(file_path))

        m = re.match(r'bytes=(\d+)-(\d*)', range_header)
        if m:
            start = int(m.group(1))
            end = m.group(2)
            end = int(end) if end else file_size - 1
        else:
            start = 0
            end = file_size - 1

        if start >= file_size:
            return Response(status=416)

        length = end - start + 1
        with open(file_path, 'rb') as f:
            f.seek(start)
            data = f.read(length)

        mime_type, _ = mimetypes.guess_type(str(file_path))
        mime_type = mime_type or 'application/octet-stream'

        rv = Response(data, 206, mimetype=mime_type, direct_passthrough=True)
        rv.headers.add('Content-Range', f'bytes {start}-{end}/{file_size}')
        rv.headers.add('Accept-Ranges', 'bytes')
        rv.headers.add('Content-Length', str(length))
        return rv


    @app.route('/video/<path:filename>')
    def video(filename):
        safe_path = (VIDEO_ROOT / filename).resolve()
        try:
            safe_path.relative_to(VIDEO_ROOT.resolve())
        except Exception:
            return Response('Forbidden', status=403)
        return send_file_partial(safe_path)

    server = make_server('127.0.0.1', 0, app)
    port = server.socket.getsockname()[1]
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()

    global base_url
    base_url = f"http://127.0.0.1:{port}"

    try:
        api = Api()
        webview.create_window(base_url, url=base_url + '/', width=1200, height=760, js_api=api)

        use_cef = False
        try:
            import cefpython3  # type: ignore
            use_cef = True
        except Exception:
            use_cef = False

        if use_cef:
            webview.start(gui='cef')
        else:
            webview.start()
    finally:
        server.shutdown()


if __name__ == "__main__":
    main()
