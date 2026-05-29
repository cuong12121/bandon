import json
from datetime import datetime
from pathlib import Path
import webbrowser
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
EXCEL_ROOT = BASE_DIR / "excel"
PAGE_SIZE = 20
VIDEO_ROOT = BASE_DIR / "video"


def resolve_video_path(raw_path):
    video_path = Path(str(raw_path).strip())
    if not video_path.is_absolute():
        video_path = (BASE_DIR / video_path).resolve()
    return video_path


def parse_close_time(value):
    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return datetime.min

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return datetime.min


def get_today_excel_path():
    now = datetime.now()
    daily_dir = EXCEL_ROOT / now.strftime("%Y") / now.strftime("%m") / now.strftime("%d")
    return daily_dir / f"{now.strftime('%Y%m%d')}.xlsx"


def load_today_orders():
    excel_path = get_today_excel_path()
    if not excel_path.exists():
        return [], excel_path

    workbook = load_workbook(excel_path, data_only=True)
    worksheet = workbook.active

    rows = []
    for row in worksheet.iter_rows(min_row=2, max_col=3, values_only=True):
        if not any(row):
            continue

        close_time = row[0] or ""
        barcode = row[1] or ""
        video_raw = row[2] or ""
        sort_key = parse_close_time(close_time)

        resolved_video = resolve_video_path(video_raw)
        exists = resolved_video.exists()
        rows.append({
            "close_time": str(close_time),
            "barcode": str(barcode),
            "video_path": str(resolved_video),
            "exists": exists,
            "sort_key": sort_key.isoformat(),
        })

    rows.sort(key=lambda item: item["sort_key"], reverse=True)
    for item in rows:
        item.pop("sort_key", None)

    # add playback URI and cleaned display link
    for item in rows:
        if item.get('exists'):
            try:
                p = Path(item['video_path'])
                item['video_uri'] = p.as_uri()
                try:
                    rel = p.relative_to(BASE_DIR)
                    item['video_link'] = str(rel).replace('\\', '/')
                except Exception:
                    item['video_link'] = p.as_posix()
            except Exception:
                item['video_uri'] = ""
                item['video_link'] = ""
        else:
            item['video_uri'] = ""
            item['video_link'] = ""

    return rows, excel_path


def build_html(rows, excel_path):
                data_json = json.dumps(rows)
                template = """<!doctype html>
<html>
<head>
        <meta charset="utf-8" />
        <title>Danh sách đơn hôm nay</title>
        <meta name="viewport" content="width=device-width,initial-scale=1" />
        <style>
                :root{--bg:#f6f8fa;--card:#ffffff;--muted:#666;--accent:#2563eb;--accent-600:#1e40af}
                html,body{height:100%;margin:0;font-family:Segoe UI, Tahoma, sans-serif;background:var(--bg);color:#111}
                .wrap{max-width:1100px;margin:20px auto;padding:20px}
                .card{background:var(--card);border-radius:10px;padding:18px;box-shadow:0 6px 18px rgba(20,20,30,0.06)}
                h1{margin:0 0 8px;font-size:20px}
                .meta{color:var(--muted);font-size:13px;margin-bottom:12px}
                table.orders{width:100%;border-collapse:collapse;background:transparent}
                table.orders th, table.orders td{padding:10px 12px;text-align:left;border-bottom:1px solid #eef2f6}
                table.orders thead th{font-weight:600;background:transparent;color:#111}
                table.orders tbody tr:nth-child(even){background:rgba(37,99,235,0.03)}
                table.orders tbody tr:hover{background:rgba(37,99,235,0.06)}
                .actions button{background:var(--accent);color:#fff;border:none;padding:8px 10px;border-radius:6px;cursor:pointer}
                .actions button:active{transform:translateY(1px)}
                .video-wrap{margin-top:14px;background:#000;padding:8px;border-radius:8px}
                #player{width:100%;max-height:480px;border-radius:6px;background:#000}
                @media (max-width:700px){table.orders th:nth-child(4), table.orders td:nth-child(4){display:none}}
        </style>
</head>
<body>
<div class="wrap">
    <div class="card">
        <h1>Danh sách đơn hôm nay</h1>
        <div class="meta">Excel: __EXCEL_PATH__</div>

        <div style="overflow:auto">
            <table class="orders" aria-label="Danh sách đơn">
                <thead><tr><th>Thời gian đóng</th><th>Mã vạch</th><th>Video</th><th>Link Video</th></tr></thead>
                <tbody id="rows"></tbody>
            </table>
        </div>

        <div class="video-wrap">
            <video id="player" controls></video>
        </div>
    </div>
</div>

<script>
const data = __DATA_JSON__;
const rowsEl = document.getElementById('rows');
const player = document.getElementById('player');
function playVideo(item){
        if(!item.video_uri){ player.removeAttribute('src'); player.load(); return; }
        player.src = item.video_uri; player.load();
}
function playVideoIndex(i){ playVideo(data[i]); }
rowsEl.innerHTML = data.map((item, i) => '<tr>' +
        '<td>' + item.close_time + '</td>' +
        '<td>' + item.barcode + '</td>' +
        '<td class="actions">' + (item.exists ? '<button onclick="playVideoIndex(' + i + ')">Xem</button>' : 'N/A') + '</td>' +
        '<td>' + item.video_link + '</td>' +
        '</tr>').join('');
</script>
</body>
</html>"""

                return template.replace('__DATA_JSON__', data_json).replace('__EXCEL_PATH__', str(excel_path))


def main():
    rows, excel_path = load_today_orders()
    html = build_html(rows, excel_path)
    out = BASE_DIR / 'list_don.html'
    out.write_text(html, encoding='utf-8')
    webbrowser.open(out.as_uri())


if __name__ == '__main__':
    main()
