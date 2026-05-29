import json
from datetime import datetime
from pathlib import Path

import webview
from openpyxl import load_workbook
import threading
import urllib.parse
import re
import mimetypes
from flask import Flask, request, Response, send_file
from werkzeug.serving import make_server

BASE_DIR = Path(__file__).resolve().parent
EXCEL_ROOT = BASE_DIR / "excel"
PAGE_SIZE = 20
VIDEO_ROOT = BASE_DIR / "video"


def get_today_excel_path():
    now = datetime.now()
    daily_dir = EXCEL_ROOT / now.strftime("%Y") / now.strftime("%m") / now.strftime("%d")
    return daily_dir / f"{now.strftime('%Y%m%d')}.xlsx"


def resolve_video_path(raw_path):
    video_path = Path(str(raw_path).strip())
    if not video_path.is_absolute():
        video_path = (BASE_DIR / video_path).resolve()
    return video_path


def parse_close_time(value):
    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return datetime.min

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return datetime.min


def load_today_orders():
    excel_path = get_today_excel_path()
    if not excel_path.exists():
        return [], excel_path

    workbook = load_workbook(excel_path, data_only=True)
    worksheet = workbook.active

    rows = []
    for row in worksheet.iter_rows(min_row=2, max_col=3, values_only=True):
        if not any(row):
            continue

        close_time = row[0] or ""
        barcode = row[1] or ""
        video_raw = row[2] or ""
        sort_key = parse_close_time(close_time)

        resolved_video = resolve_video_path(video_raw)
        exists = resolved_video.exists()
        rows.append({
          "close_time": str(close_time),
          "barcode": str(barcode),
          "video_path": str(resolved_video),
          "exists": exists,
          "sort_key": sort_key.isoformat(),
        })

    rows.sort(key=lambda item: item["sort_key"], reverse=True)
    for item in rows:
        item.pop("sort_key", None)

    return rows, excel_path


def build_html(rows, excel_path):
    data_json = json.dumps(rows)

    return f"""
<!doctype html>
<html>
<head>
  <meta charset=\"utf-8\" />
  <title>Danh sach don hom nay</title>
  <style>
    :root {{
      --bg: #f4efe4;
      --surface: #fff8ee;
      --ink: #1f2328;
      --muted: #5f676f;
      --accent: #005f73;
      --accent-2: #ee9b00;
      --line: #d9cbb0;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", Tahoma, sans-serif;
      color: var(--ink);
      background: radial-gradient(circle at 15% 10%, #fff5dd 0%, var(--bg) 45%);
      min-height: 100vh;
    }}
    .wrap {{
      max-width: 1100px;
      margin: 24px auto;
      padding: 0 16px 24px;
    }}
    .head {{
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: 14px;
      padding: 14px 16px;
      margin-bottom: 14px;
    }}
    .title {{
      margin: 0;
      font-size: 24px;
      color: var(--accent);
    }}
    .meta {{
      margin-top: 6px;
      color: var(--muted);
      font-size: 13px;
      word-break: break-all;
    }}
    .grid {{
      display: grid;
      grid-template-columns: 1fr;
      gap: 14px;
    }}
    @media (min-width: 980px) {{
      .grid {{ grid-template-columns: 1.2fr .8fr; }}
    }}
    .card {{
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: 14px;
      padding: 12px;
      box-shadow: 0 8px 20px rgba(31, 35, 40, 0.06);
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 14px;
    }}
    th, td {{
      border-bottom: 1px solid var(--line);
      text-align: left;
      padding: 10px 8px;
      vertical-align: top;
    }}
    th {{
      color: var(--accent);
      font-weight: 700;
      background: #fff2dc;
      position: sticky;
      top: 0;
    }}
    .table-wrap {{
      max-height: 66vh;
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 10px;
      background: #fff;
    }}
    .btn {{
      border: 0;
      border-radius: 8px;
      padding: 7px 10px;
      background: var(--accent);
      color: #fff;
      cursor: pointer;
      font-weight: 600;
    }}
    .btn:disabled {{
      background: #8a8f95;
      cursor: not-allowed;
    }}
    .pager {{
      margin-top: 10px;
      display: flex;
      align-items: center;
      gap: 8px;
    }}
    .pager-info {{
      color: var(--muted);
      font-size: 13px;
      margin-left: 4px;
    }}
    .video-box {{
      border: 1px dashed var(--line);
      border-radius: 10px;
      padding: 10px;
      min-height: 240px;
      background: #fff;
    }}
    .video-title {{
      margin: 0 0 8px;
      color: var(--accent);
      font-size: 15px;
    }}
    #videoInfo {{
      font-size: 12px;
      color: var(--muted);
      word-break: break-all;
      margin: 0 0 8px;
    }}
    video {{
      width: 100%;
      max-height: 420px;
      background: #000;
      border-radius: 8px;
    }}
    .empty {{
      color: var(--muted);
      font-style: italic;
    }}
  </style>
</head>
<body>
  <div class=\"wrap\">
    <div class=\"head\">
      <h1 class=\"title\">Danh sach don hom nay</h1>
      <div class=\"meta\">Excel: {excel_path}</div>
      <div class=\"meta\" id=\"count\"></div>
    </div>

    <div class=\"grid\">
      <section class=\"card\">
        <div class=\"table-wrap\">
          <table>
            <thead>
              <tr>
                <th>Thoi gian dong</th>
                <th>Ma vach</th>
                <th>Video</th>
              </tr>
            </thead>
            <tbody id=\"rows\"></tbody>
          </table>
        </div>
        <div class=\"pager\">
          <button class=\"btn\" id=\"prevBtn\" onclick=\"prevPage()\">Trang truoc</button>
          <button class=\"btn\" id=\"nextBtn\" onclick=\"nextPage()\">Trang sau</button>
          <span class=\"pager-info\" id=\"pageInfo\">Trang 1/1</span>
        </div>
      </section>

      <section class=\"card\">
        <div class=\"video-box\">
          <h3 class=\"video-title\">Xem video don</h3>
          <p id=\"videoInfo\">Chon mot don de xem video.</p>
          <video id=\"player\" controls></video>
        </div>
      </section>
    </div>
  </div>

  <script>
    const data = {data_json};
    const pageSize = {PAGE_SIZE};
    let currentPage = 1;
    const rowsEl = document.getElementById('rows');
    const countEl = document.getElementById('count');
    const player = document.getElementById('player');
    const info = document.getElementById('videoInfo');
    const pageInfoEl = document.getElementById('pageInfo');
    const prevBtn = document.getElementById('prevBtn');
    const nextBtn = document.getElementById('nextBtn');

    function playVideo(item) {{
      if (!item.video_uri) {{
        player.removeAttribute('src');
        player.load();
        info.textContent = 'Khong tim thay file video: ' + item.video_path;
        return;
      }}

      player.src = item.video_uri;
      player.load();
      info.textContent = item.video_path;
    }}

    function renderTable() {{
      if (data.length === 0) {{
        rowsEl.innerHTML = '<tr><td colspan="3" class="empty">Khong co du lieu don cho ngay hom nay.</td></tr>';
        countEl.textContent = 'Tong don: 0';
        pageInfoEl.textContent = 'Trang 0/0';
        prevBtn.disabled = true;
        nextBtn.disabled = true;
        return;
      }}

      const totalPages = Math.ceil(data.length / pageSize);
      if (currentPage > totalPages) {{
        currentPage = totalPages;
      }}

      const start = (currentPage - 1) * pageSize;
      const end = start + pageSize;
      const pageRows = data.slice(start, end);

      countEl.textContent = 'Tong don: ' + data.length;
      pageInfoEl.textContent = 'Trang ' + currentPage + '/' + totalPages;
      prevBtn.disabled = currentPage <= 1;
      nextBtn.disabled = currentPage >= totalPages;

      rowsEl.innerHTML = pageRows.map((item, index) => {{
        const actualIndex = start + index;
        const disabled = item.exists ? '' : 'disabled';
        const btn = '<button class="btn" ' + disabled + ' onclick="playVideo(data[' + actualIndex + '])">Xem</button>';

        return '<tr>' +
          '<td>' + item.close_time + '</td>' +
          '<td>' + item.barcode + '</td>' +
          '<td>' + btn + '</td>' +
          '</tr>';
      }}).join('');
    }}

    function nextPage() {{
      const totalPages = Math.ceil(data.length / pageSize);
      if (currentPage < totalPages) {{
        currentPage += 1;
        renderTable();
      }}
    }}

    function prevPage() {{
      if (currentPage > 1) {{
        currentPage -= 1;
        renderTable();
      }}
    }}

    renderTable();
  </script>
</body>
</html>
"""


def main():
  # start tiny HTTP server to serve video files from VIDEO_ROOT
  VIDEO_ROOT.mkdir(parents=True, exist_ok=True)

  app = Flask(__name__)

  rows, excel_path = load_today_orders()

  # convert video paths to flask URLs served by /video/<path:filename>
  for item in rows:
    if item.get('exists'):
      try:
        # Use the original file URI (file://) for video source instead of the local HTTP URL
        item['video_uri'] = Path(item['video_path']).as_uri()
      except Exception:
        item['video_uri'] = ""
    else:
      item['video_uri'] = ""

  html = build_html(rows, excel_path)

  @app.route('/')
  def index():
    return html

  def send_file_partial(path):
    file_path = Path(path)
    if not file_path.exists():
      return Response('File not found', status=404)

    file_size = file_path.stat().st_size
    range_header = request.headers.get('Range', None)
    if not range_header:
      return send_file(str(file_path))

    m = re.match(r'bytes=(\d+)-(\d*)', range_header)
    if m:
      start = int(m.group(1))
      end = m.group(2)
      end = int(end) if end else file_size - 1
    else:
      start = 0
      end = file_size - 1

    if start >= file_size:
      return Response(status=416)

    length = end - start + 1
    with open(file_path, 'rb') as f:
      f.seek(start)
      data = f.read(length)

    mime_type, _ = mimetypes.guess_type(str(file_path))
    mime_type = mime_type or 'application/octet-stream'

    rv = Response(data, 206, mimetype=mime_type, direct_passthrough=True)
    rv.headers.add('Content-Range', f'bytes {start}-{end}/{file_size}')
    rv.headers.add('Accept-Ranges', 'bytes')
    rv.headers.add('Content-Length', str(length))
    return rv

  @app.route('/video/<path:filename>')
  def video(filename):
    safe_path = (VIDEO_ROOT / filename).resolve()
    try:
      # ensure file is under VIDEO_ROOT
      safe_path.relative_to(VIDEO_ROOT.resolve())
    except Exception:
      return Response('Forbidden', status=403)
    return send_file_partial(safe_path)

  # start flask server on an ephemeral port in background
  server = make_server('127.0.0.1', 0, app)
  port = server.socket.getsockname()[1]
  thread = threading.Thread(target=server.serve_forever, daemon=True)
  thread.start()

  base_url = f"http://127.0.0.1:{port}"

  try:
    webview.create_window(base_url, url=base_url + '/', width=1200, height=760)
    webview.start()
  finally:
    server.shutdown()


if __name__ == "__main__":
    main()